import openpyxl
from database.database import SessionLocal
from models.department import Department
from models.asset import Asset
from models.technician import Technician
from models.requester import Requester
from models.task_request import TaskRequest
from models.work_order import WorkOrder

DATA_PATH = "RTknits CMMS Data Pack/RTknits CMMS Data Pack/" # adjust if your folder name differs

db = SessionLocal()

print("Seeding departments and users...")
wb = openpyxl.load_workbook(DATA_PATH + "users_department.xlsx")
ws = wb.active
dept_map = {}  # dept name -> dept_id

for row in list(ws.iter_rows(min_row=2, values_only=True)):
    user_id, dept_name, location, username = row
    if not dept_name:
        continue
    dept_name = str(dept_name).strip()
    if dept_name not in dept_map:
        existing = db.query(Department).filter_by(name=dept_name).first()
        if not existing:
            dept = Department(name=dept_name, location=location)
            db.add(dept)
            db.flush()
            dept_map[dept_name] = dept.dept_id
        else:
            dept_map[dept_name] = existing.dept_id

db.commit()
print(f"  {len(dept_map)} departments loaded")

print("Seeding assets...")
wb = openpyxl.load_workbook(DATA_PATH + "Assets.xlsx")
ws = wb.active
asset_map = {}  # original asset ID -> new asset_id

for row in list(ws.iter_rows(min_row=2, values_only=True)):
    orig_id = row[0]
    category = str(row[1]).strip() if row[1] else None
    dept_name = str(row[4]).strip() if row[4] else None
    location = str(row[6]).strip() if row[6] else None
    name = str(row[9]).strip() if row[9] else f"Asset {orig_id}"
    status = str(row[15]).strip() if row[15] else None

    dept_id = dept_map.get(dept_name) if dept_name else None

    asset = Asset(
        name=name,
        category=category,
        dept_id=dept_id
    )
    db.add(asset)
    db.flush()
    asset_map[orig_id] = asset.asset_id

db.commit()
print(f"  {len(asset_map)} assets loaded")

print("Seeding technicians...")
wb = openpyxl.load_workbook(DATA_PATH + "Technicians.xlsx")
ws = wb.active
tech_map = {}  # name -> tech_id

for row in list(ws.iter_rows(min_row=2, values_only=True)):
    orig_id, salary, flag_clock, flag_pic, job_title, nationality, prox, team_leader, location, name = row
    if not name:
        continue
    name = str(name).strip()
    tech = Technician(
        name=name,
        trade=str(job_title).strip() if job_title else None,
        pool=str(location).strip() if location else None,
        phone_number=None,
        on_shift=False
    )
    db.add(tech)
    db.flush()
    tech_map[name] = tech.tech_id

db.commit()
print(f"  {len(tech_map)} technicians loaded")

print("Seeding task requests...")
wb = openpyxl.load_workbook(DATA_PATH + "Tasks.xlsx")
ws = wb.active
task_map = {}  # original task ID -> new request_id

for row in list(ws.iter_rows(min_row=2, values_only=True)):
    orig_id = row[0]
    orig_asset_id = row[1]
    description = str(row[4]).strip() if row[4] else None
    created_at = row[8]

    asset_id = asset_map.get(orig_asset_id) if orig_asset_id else None

    task = TaskRequest(
        asset_id=asset_id,
        raw_text=description,
    )
    db.add(task)
    db.flush()
    task_map[orig_id] = task.request_id

db.commit()
print(f"  {len(task_map)} task requests loaded")

print("Seeding work orders...")
wb = openpyxl.load_workbook(DATA_PATH + "Workorder.xlsx")
ws = wb.active
count = 0

for row in list(ws.iter_rows(min_row=2, values_only=True)):
    orig_wo_id = row[0]
    orig_task_id = row[1]
    description = str(row[5]).strip() if row[5] else None
    status = str(row[14]).strip() if row[14] else "open"

    request_id = task_map.get(orig_task_id) if orig_task_id else None

    wo = WorkOrder(
        request_id=request_id,
        priority="P2",  # historical data, default to P2
        status=status.lower() if status else "open",
        description=description
    )
    db.add(wo)
    count += 1

db.commit()
print(f"  {count} work orders loaded")

db.close()
print("\nAll done! Database seeded with real RT Knits data.")